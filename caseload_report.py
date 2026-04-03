"""
Caseload Report Automation Script
St. Vincent de Paul CARES — Data Systems Team

Replaces the manual process of combining CaseWorthy Excel exports
into a formatted, multi-tab Caseload Report workbook.

Usage:
    python caseload_report.py [-o output.xlsx]

Place input files in these folders (script auto-detects .xlsx files):
    input/data_report_card/    — Data Report Card export
    input/legal_referral/      — Legal Services Referral export
"""

import argparse
import csv
import glob
import logging
import os
import sys
from collections import OrderedDict
from datetime import date, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CONFIGURABLE CONSTANTS
# ---------------------------------------------------------------------------

# Input folder paths (relative to script directory)
INPUT_DIR_DATA_REPORT_CARD = os.path.join("input", "data_report_card")
INPUT_DIR_LEGAL_REFERRAL = os.path.join("input", "legal_referral")
CONFIG_DIR_STAFF_ROSTER = os.path.join("config", "staff_roster")
CONFIG_SITE_TAB_MAPPING = os.path.join("config", "site_tab_mapping.csv")
CONFIG_PROGRAM_VALIDATION = os.path.join("config", "program_validation.xlsx")
OUTPUT_DIR = "output"

# Column renames applied to the Data Report Card
COLUMN_RENAMES = {
    "Case Manager": "Assigned Staff",
    "Receiving Shallow Subsidy": "Current Receive ShallowSub",
    "Referred From HUD-VASH": "Referred From HUDVASH",
}

# Columns kept from the Data Report Card after dropping unused ones (Step 3).
# These are the 15 direct columns that survive into the final output.
# "Last Case Note Date Per Prog" is used to derive Days With no Service/Contact.
# "Current Office Location" is extracted separately for the MidFlorida tab.
KEEP_COLUMNS_FROM_DRC = [
    "Client ID",
    "First Name",
    "Last Name",
    "# Enrolled Family Members",
    "Program Name",
    "Begin Date",
    "Days Enrolled",
    "Move-In Date",
    "Assigned Staff",
    "Staff Job Type",
    "Days since Last Recert/Update",
    "Current Receive ShallowSub",
    "Referred From HUDVASH",
    "Connection With SOAR",
    "Last Case Note Date Per Prog",
    "PQI Review",
    "Peer Review",
]

# The 20 output columns in exact order for the "All" tab
ALL_COLUMNS_ORDERED = [
    "Client ID",
    "First Name",
    "Last Name",
    "# Enrolled Family Members",
    "Program Name",
    "Begin Date",
    "Days Enrolled",
    "Move-In Date",
    "Assigned Staff",
    "Staff Job Type",
    "Last 90 Day Recert",
    "Days since Last Recert/Update",
    "Current Receive ShallowSub",
    "Referred From HUDVASH",
    "Connection With SOAR",
    "Received Legal Assistance",
    "Days With no Service/Contact",
    "Housed Not Housed",
    "PQI Review",
    "Peer Review",
]

# Critical columns that MUST exist in the Data Report Card
CRITICAL_DRC_COLUMNS = [
    "Event",
    "Client ID",
    "Program Name",
    "Case Manager",
    "Last Case Note Date Per Prog",
]

# Non-SSVF programs — these get N/A for SOAR, ShallowSub, HUDVASH, and Recert.
# Update this list as new non-SSVF programs are added to CaseWorthy.
NON_SSVF_PROGRAMS = [
    "All-County-VA-Suicide-Prevention 1114",
    "Charlotte County-SHIP-RRH 1305",
    "Tampa-THHI-CDBG DAP CES 1107",
]

# Keywords that indicate a program IS SSVF-funded (used for heuristic warnings)
SSVF_KEYWORDS = ["SSVF", "EHA", "GPD", "HCHV", "HUD-CoC", "CoC", "ESG", "PSH"]

# Tab order for the output workbook
SITE_TAB_ORDER = [
    "All",
    "Charlotte",
    "Charlotte Shelter",
    "Citrus",
    "FOX",
    "GPD",
    "Lake",
    "Orlando",
    "Pasco",
    "Pinellas",
    "Polk",
    "PSH",
    "San Juan",
    "Sarasota",
    "Sebring",
    "SouthWest",
    "Tampa",
]

# Tabs that get a Location column inserted at position 10 (from Current Office Location)
LOCATION_TABS = {"Citrus", "Lake"}

# Columns dropped from the FOX tab (reduced 12-column layout)
FOX_DROP_COLUMNS = [
    "Move-In Date",
    "Last 90 Day Recert",
    "Days since Last Recert/Update",
    "Current Receive ShallowSub",
    "Referred From HUDVASH",
    "Connection With SOAR",
    "Housed Not Housed",
]

# Formatting constants
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_ROW_FILL = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(color="9C0006")
DATE_FORMAT = "MM/DD/YYYY"
DATE_COLUMNS = {"Begin Date", "Move-In Date", "Last 90 Day Recert"}
CENTER_COLUMNS = {"# Enrolled Family Members"}
RED_THRESHOLD_COLUMNS = {"Days With no Service/Contact"}
RED_THRESHOLD_DAYS = 30


# ---------------------------------------------------------------------------
# FILE DISCOVERY
# ---------------------------------------------------------------------------


def find_xlsx_in_folder(folder_path):
    """Find a single .xlsx file in the given folder.

    Args:
        folder_path: Path to the input folder

    Returns:
        Path to the .xlsx file

    Raises:
        FileNotFoundError: If no .xlsx files found
        ValueError: If multiple .xlsx files found
    """
    # Resolve relative to script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    abs_folder = os.path.join(script_dir, folder_path)

    if not os.path.isdir(abs_folder):
        raise FileNotFoundError(
            f"Input folder not found: {abs_folder}\n"
            f"Create it and place your .xlsx export inside."
        )

    xlsx_files = glob.glob(os.path.join(abs_folder, "*.xlsx"))
    # Exclude temp files (Excel lock files start with ~$)
    xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith("~$")]

    if not xlsx_files:
        raise FileNotFoundError(
            f"No .xlsx files found in {abs_folder}\n"
            f"Place your CaseWorthy export in this folder."
        )

    if len(xlsx_files) > 1:
        raise ValueError(
            f"Multiple .xlsx files found in {abs_folder}:\n"
            + "\n".join(f"  - {os.path.basename(f)}" for f in xlsx_files)
            + "\nPlease keep only the file you want to process."
        )

    logger.info("Found: %s", xlsx_files[0])
    return xlsx_files[0]


# ---------------------------------------------------------------------------
# LOADER FUNCTIONS
# ---------------------------------------------------------------------------


def load_staff_roster():
    """Load the staff roster from config/staff_roster/ if available.

    Maps Login Name → (Last Name, First Name, Job Type) for resolving
    abbreviated staff names in the Data Report Card.

    Returns:
        dict or None: {login_name: {"full_name": "Last, First", "job_type": "..."}}
        Returns None if no roster file is found.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    roster_dir = os.path.join(script_dir, CONFIG_DIR_STAFF_ROSTER)

    if not os.path.isdir(roster_dir):
        logger.info("No staff roster folder found — using raw staff names")
        return None

    xlsx_files = glob.glob(os.path.join(roster_dir, "*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not os.path.basename(f).startswith("~$")]

    if not xlsx_files:
        logger.info("No staff roster file found in %s — using raw staff names", roster_dir)
        return None

    if len(xlsx_files) > 1:
        logger.warning("Multiple files in %s — using first: %s", roster_dir, xlsx_files[0])

    filepath = xlsx_files[0]
    logger.info("Reading staff roster: %s", filepath)
    df = pd.read_excel(filepath, engine="openpyxl")
    df.columns = df.columns.str.strip()

    required = ["Login Name", "Last Name", "First Name"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        logger.warning("Staff roster missing columns %s — skipping staff mapping", missing)
        return None

    roster = {}
    for _, row in df.iterrows():
        login = str(row["Login Name"]).strip()
        if not login or login == "nan":
            continue
        last = str(row.get("Last Name", "")).strip()
        first = str(row.get("First Name", "")).strip()
        job_type = str(row.get("Job Type", "")).strip() if "Job Type" in df.columns else ""
        if job_type == "nan":
            job_type = ""
        roster[login] = {
            "full_name": f"{last}, {first}" if last and first else login,
            "job_type": job_type,
        }

    logger.info("  Loaded %d staff entries", len(roster))
    return roster


def load_site_tab_mapping():
    """Load site tab mapping from config/site_tab_mapping.csv if available.

    Returns:
        list of dicts or None: Each dict has keys: tab_name, match_type, match_value.
        Returns None if the file doesn't exist (falls back to hardcoded filters).
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(script_dir, CONFIG_SITE_TAB_MAPPING)

    if not os.path.isfile(csv_path):
        return None

    rows = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    logger.info("Loaded site tab mapping: %d rules from %s", len(rows), csv_path)
    return rows


def load_program_validation():
    """Load per-program, per-field validation rules from config/program_validation.xlsx.

    Each row specifies a program (exact or contains match) and which fields
    should be validated (Yes) or blanked out (No) for that program.

    Validation columns: Connection With SOAR, Current Receive ShallowSub,
    Referred From HUDVASH, Last 90 Day Recert, Received Legal Assistance,
    Housed Not Housed.

    Returns:
        list of dicts, each with keys:
            "program": str (program name or pattern)
            "match_type": "exact" or "contains"
            "fields": dict of {field_name: bool} where False = blank out
        Returns None if the file doesn't exist (falls back to hardcoded logic).
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(script_dir, CONFIG_PROGRAM_VALIDATION)

    if not os.path.isfile(xlsx_path):
        logger.info("No program_validation.xlsx found — using hardcoded defaults")
        return None

    logger.info("Reading program validation: %s", xlsx_path)
    df = pd.read_excel(xlsx_path, engine="openpyxl")
    df.columns = df.columns.str.strip()

    if "Program Name" not in df.columns:
        logger.warning("program_validation.xlsx missing 'Program Name' column — skipping")
        return None

    field_columns = [
        "Connection With SOAR",
        "Current Receive ShallowSub",
        "Referred From HUDVASH",
        "Last 90 Day Recert",
        "Received Legal Assistance",
        "Housed Not Housed",
    ]

    rules = []
    for _, row in df.iterrows():
        prog = str(row["Program Name"]).strip()
        if not prog or prog == "nan":
            continue

        match_type = str(row.get("Match Type", "exact")).strip().lower()
        if match_type == "nan" or not match_type:
            match_type = "exact"

        fields = {}
        for col in field_columns:
            if col in df.columns:
                val = str(row.get(col, "")).strip().lower()
                # "No" = blank this field. Blank/empty/anything else = pass through data
                if val in ("no", "n", "false", "0"):
                    fields[col] = False
                # else: not in fields dict = no rule, pass through
            # Column not in Excel = no rule, pass through

        rules.append({
            "program": prog,
            "match_type": match_type,
            "fields": fields,
        })

    logger.info("  Loaded %d program validation rules", len(rules))
    return rules


def load_data_report_card(filepath):
    """Load and validate the Data Report Card export.

    Returns:
        tuple: (DataFrame with all columns, Series of Current Office Location)
    """
    logger.info("Reading Data Report Card: %s", filepath)
    df = pd.read_excel(filepath, engine="openpyxl")
    df.columns = df.columns.str.strip()  # Clean whitespace from column names
    logger.info("  Loaded %d rows, %d columns", len(df), len(df.columns))

    # Validate critical columns
    missing = [c for c in CRITICAL_DRC_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"Data Report Card is missing critical columns: {missing}. "
            f"Found columns: {list(df.columns)}"
        )

    # Extract Current Office Location before any processing (for MidFlorida tab)
    if "Current Office Location" in df.columns:
        office_location = df["Current Office Location"].copy()
    else:
        logger.warning("'Current Office Location' column not found — MidFlorida Location will be blank")
        office_location = pd.Series("", index=df.index)

    return df, office_location


def load_legal_referral(filepath):
    """Load the Legal Services Referral report.

    Filters for Approved referrals, converts Client ID to integer,
    and relabels status as 'Received'.

    Returns:
        DataFrame with columns: CW Client ID, Referral Status
    """
    logger.info("Reading Legal Services Referral: %s", filepath)
    df = pd.read_excel(filepath, engine="openpyxl")
    df.columns = df.columns.str.strip()  # Clean whitespace from column names
    logger.info("  Loaded %d rows", len(df))

    # Validate required columns
    required = ["CW Client ID", "Referral Status"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Legal Services Referral missing columns: {missing}")

    # Keep only needed columns
    df = df[required].copy()

    # Convert Client ID to numeric
    df["CW Client ID"] = pd.to_numeric(df["CW Client ID"], errors="coerce")
    df = df.dropna(subset=["CW Client ID"])
    df["CW Client ID"] = df["CW Client ID"].astype("Int64")

    # Filter for Approved only, then relabel
    df = df[df["Referral Status"] == "Approved"].copy()
    df["Referral Status"] = "Received"

    # Deduplicate — one marker per client is enough
    df = df.drop_duplicates(subset=["CW Client ID"], keep="first")

    logger.info("  %d clients with approved legal referrals", len(df))
    return df


# ---------------------------------------------------------------------------
# PROCESSING FUNCTIONS
# ---------------------------------------------------------------------------


def process_main_sheet(drc, office_location, legal, staff_roster=None, program_validation=None):
    """Process the main 'All' sheet.

    Args:
        drc: Raw Data Report Card DataFrame
        office_location: Series of Current Office Location values (aligned to drc index)
        legal: Cleaned Legal Services Referral DataFrame
        staff_roster: Optional dict mapping login names to full names and job types
        program_validation: Optional list of validation rules from load_program_validation()

    Returns:
        tuple: (Processed All-tab DataFrame, aligned office_location Series)
    """
    df = drc.copy()

    # ── Step 1: Rename columns and replace Event values ──
    df.rename(columns=COLUMN_RENAMES, inplace=True)
    df["Event"] = df["Event"].replace("At Exit", "ZAT Exit")

    # ── Step 1b: Map staff login names to full names + job type ──
    if staff_roster and "Assigned Staff" in df.columns:
        def resolve_staff(login_val):
            if pd.isna(login_val):
                return login_val
            login = str(login_val).strip()
            if login in staff_roster:
                return staff_roster[login]["full_name"]
            return login_val

        def resolve_job_type(login_val):
            if pd.isna(login_val):
                return ""
            login = str(login_val).strip()
            if login in staff_roster:
                return staff_roster[login]["job_type"]
            return ""

        # Resolve before overwriting — need original login names for lookup
        df["Staff Job Type"] = df["Assigned Staff"].apply(resolve_job_type)
        df["Assigned Staff"] = df["Assigned Staff"].apply(resolve_staff)

        matched = (df["Staff Job Type"] != "").sum()
        logger.info("Staff roster: %d/%d staff names resolved", matched, len(df))
    else:
        df["Staff Job Type"] = ""

    # ── Step 2: Sort and deduplicate ──
    df.sort_values(by=["Event", "Client ID", "Program Name"], inplace=True)
    before_count = len(df)
    df.drop_duplicates(subset=["Client ID", "Program Name"], keep="first", inplace=True)
    df.reset_index(drop=False, inplace=True)  # preserve original index for office_location alignment
    logger.info(
        "Dedup: %d -> %d rows (%d duplicates removed)",
        before_count,
        len(df),
        before_count - len(df),
    )

    # Align office_location to the surviving rows
    if "index" in df.columns:
        office_loc_aligned = office_location.reindex(df["index"]).reset_index(drop=True)
        df.drop(columns=["index"], inplace=True)
    else:
        office_loc_aligned = pd.Series("", index=df.index)

    # ── Step 3: Drop unused columns ──
    available_keep = [c for c in KEEP_COLUMNS_FROM_DRC if c in df.columns]
    missing_keep = [c for c in KEEP_COLUMNS_FROM_DRC if c not in df.columns]
    if missing_keep:
        logger.warning("Expected columns not found in Data Report Card: %s", missing_keep)
        logger.info("  Available columns after rename: %s", list(df.columns))
    df = df[available_keep].copy()
    df.reset_index(drop=True, inplace=True)

    # ── Step 4: Calculate Last 90 Day Recert ──
    today = date.today()

    if "Days since Last Recert/Update" in df.columns:
        def calc_recert(days_val):
            if pd.isna(days_val):
                return pd.NaT
            try:
                return today - timedelta(days=int(days_val))
            except (ValueError, TypeError):
                return pd.NaT

        df["Last 90 Day Recert"] = df["Days since Last Recert/Update"].apply(calc_recert)

        # Insert at position 10 (after Staff Job Type, before Days since Last Recert/Update)
        cols = list(df.columns)
        cols.remove("Last 90 Day Recert")
        cols.insert(10, "Last 90 Day Recert")
        df = df[cols]
    else:
        logger.warning("'Days since Last Recert/Update' not found — Last 90 Day Recert will be blank")
        df["Last 90 Day Recert"] = pd.NaT

    # ── Step 5: Apply per-program validation rules (from config) ──
    # Determines which fields to blank out on a per-program basis.
    # Fields set to "No" in program_validation.xlsx get blanked.
    pn = df["Program Name"].fillna("")

    # Map output column names to what gets blanked
    VALIDATION_FIELD_MAP = {
        "Connection With SOAR": ("Connection With SOAR", ""),
        "Current Receive ShallowSub": ("Current Receive ShallowSub", ""),
        "Referred From HUDVASH": ("Referred From HUDVASH", ""),
        "Last 90 Day Recert": ("Last 90 Day Recert", pd.NaT),
        "Received Legal Assistance": ("Received Legal Assistance", ""),
        "Housed Not Housed": ("Housed Not Housed", ""),
    }
    # Days since Last Recert/Update is blanked alongside Last 90 Day Recert

    if program_validation:
        covered_programs = set()
        total_blanked = 0

        for rule in program_validation:
            pattern = rule["program"]
            match_type = rule["match_type"]
            fields = rule["fields"]

            # Build mask for this rule
            if match_type == "contains":
                mask = pn.str.contains(pattern, case=False, na=False)
            else:
                mask = pn == pattern

            if not mask.any():
                continue

            # Track which programs are covered by validation rules
            covered_programs.update(pn[mask].unique())

            # Blank out only fields explicitly set to "No"
            # Fields not in the dict = no rule, pass through real data
            for field_name, validate in fields.items():
                if validate is not False:
                    continue
                col_name, blank_val = VALIDATION_FIELD_MAP.get(field_name, (None, None))
                if col_name and col_name in df.columns:
                    df.loc[mask, col_name] = blank_val
                    total_blanked += mask.sum()
                # Also blank Days since Last Recert/Update when Recert is No
                if field_name == "Last 90 Day Recert" and "Days since Last Recert/Update" in df.columns:
                    df.loc[mask, "Days since Last Recert/Update"] = pd.NA

        logger.info("Validation rules applied: %d programs covered", len(covered_programs))

        # Warn about programs not covered by any validation rule
        all_programs = pn.unique()
        for prog in all_programs:
            if prog in covered_programs:
                continue
            has_keyword = any(kw in str(prog) for kw in SSVF_KEYWORDS)
            if not has_keyword:
                logger.warning(
                    "Program '%s' has no validation rule in program_validation.xlsx "
                    "and does not match any SSVF keyword — verify if it needs one",
                    prog,
                )
    else:
        # Fallback: hardcoded non-SSVF behavior (blanks all SSVF fields)
        non_ssvf_mask = pn.isin(NON_SSVF_PROGRAMS)
        for col in ["Current Receive ShallowSub", "Referred From HUDVASH", "Connection With SOAR"]:
            if col in df.columns:
                df.loc[non_ssvf_mask, col] = ""
        if "Last 90 Day Recert" in df.columns:
            df.loc[non_ssvf_mask, "Last 90 Day Recert"] = pd.NaT
        if "Days since Last Recert/Update" in df.columns:
            df.loc[non_ssvf_mask, "Days since Last Recert/Update"] = pd.NA

    # ── Step 6: VLOOKUP — Received Legal Assistance ──
    legal_lookup = legal.rename(
        columns={"CW Client ID": "Client ID", "Referral Status": "Received Legal Assistance"}
    )
    df = df.merge(legal_lookup, on="Client ID", how="left")

    unmatched_legal = df["Received Legal Assistance"].isna().sum()
    matched_legal = len(df) - unmatched_legal
    logger.info("Legal referral: %d matched, %d unmatched -> N/A", matched_legal, unmatched_legal)
    df["Received Legal Assistance"] = df["Received Legal Assistance"].fillna("")

    # ── Step 7: Days With no Service/Contact (from Last Case Note Date Per Prog) ──
    if "Last Case Note Date Per Prog" in df.columns:
        df["Last Case Note Date Per Prog"] = pd.to_datetime(
            df["Last Case Note Date Per Prog"], errors="coerce"
        )

        def calc_days_no_service(note_date):
            if pd.isna(note_date):
                return pd.NA
            try:
                return (pd.Timestamp(today) - note_date).days
            except (ValueError, TypeError):
                return pd.NA

        df["Days With no Service/Contact"] = df["Last Case Note Date Per Prog"].apply(
            calc_days_no_service
        )

        has_note = df["Last Case Note Date Per Prog"].notna().sum()
        no_note = df["Last Case Note Date Per Prog"].isna().sum()
        logger.info(
            "Days With no Service/Contact: %d calculated, %d blank (no case note date)",
            has_note,
            no_note,
        )

        # Drop the intermediate column — not part of final output
        df.drop(columns=["Last Case Note Date Per Prog"], inplace=True)
    else:
        df["Days With no Service/Contact"] = pd.NA
        logger.warning("'Last Case Note Date Per Prog' column not found — Days With no Service/Contact will be blank")

    # ── Step 8: Housed / Not Housed ──
    df["Housed Not Housed"] = df["Move-In Date"].apply(
        lambda x: "Housed" if pd.notna(x) else "Not Housed"
    )
    non_rrh_mask = ~df["Program Name"].str.contains("RRH", case=False, na=False)
    df.loc[non_rrh_mask, "Housed Not Housed"] = ""

    # ── Step 9: PQI Review and Peer Review (Yes / No / blank) ──
    for col in ["PQI Review", "Peer Review"]:
        if col in df.columns:
            df[col] = df[col].apply(_normalize_yes_no)

    # ── Final column ordering ──
    final_cols = [c for c in ALL_COLUMNS_ORDERED if c in df.columns]
    df = df[final_cols]

    logger.info("All tab: %d rows, %d columns", len(df), len(df.columns))
    return df, office_loc_aligned


def _normalize_yes_no(value):
    """Normalize PQI/Peer Review values to Yes, No, or blank.

    - Blank/null → blank ("")
    - "Yes" (any case) → "Yes"
    - "No" (any case) → "No"
    - Any other non-blank value → "Yes"
    """
    if pd.isna(value) or str(value).strip() == "":
        return ""
    val = str(value).strip()
    if val.lower() == "no":
        return "No"
    if val.lower() == "yes":
        return "Yes"
    # Any other non-blank value indicates the review was done
    return "Yes"


def _build_filters_from_csv(mapping_rows, pn, ol, index):
    """Build site tab filter masks from CSV mapping rules.

    Supports match types:
        prefix            — Program Name starts with value
        contains          — Program Name contains value (case-insensitive)
        location_contains — Current Office Location contains value (case-insensitive)
        exclude_contains  — Exclude rows where Program Name contains value
        require_contains  — Require rows where Program Name contains value

    Multiple rules for the same tab are OR'd together (except exclude/require
    which are AND'd as modifiers).
    """
    # Group rules by tab name
    tab_rules = OrderedDict()
    for row in mapping_rows:
        tab = row["tab_name"]
        if tab not in tab_rules:
            tab_rules[tab] = []
        tab_rules[tab].append(row)

    filters = {}
    for tab_name, rules in tab_rules.items():
        include_mask = pd.Series(False, index=index)
        exclude_mask = pd.Series(False, index=index)
        require_mask = pd.Series(True, index=index)

        for rule in rules:
            mt = rule["match_type"]
            mv = rule["match_value"]

            if mt == "prefix":
                include_mask = include_mask | pn.str.startswith(mv)
            elif mt == "contains":
                include_mask = include_mask | pn.str.contains(mv, case=False, na=False)
            elif mt == "location_contains":
                include_mask = include_mask | ol.str.contains(mv, case=False, na=False)
            elif mt == "exclude_contains":
                exclude_mask = exclude_mask | pn.str.contains(mv, case=False, na=False)
            elif mt == "require_contains":
                require_mask = require_mask & pn.str.contains(mv, case=False, na=False)

        filters[tab_name] = include_mask & ~exclude_mask & require_mask

    return filters


def create_site_tabs(df_all, office_location, site_tab_mapping=None):
    """Create site tabs from the All sheet data.

    Args:
        df_all: Processed All-tab DataFrame
        office_location: Series of Location values aligned to df_all
        site_tab_mapping: Optional list of mapping rules from CSV config

    Returns:
        OrderedDict of {tab_name: DataFrame} in SITE_TAB_ORDER
    """
    pn = df_all["Program Name"].fillna("")
    ol = office_location.fillna("")

    # Build filter masks — from CSV config if available, otherwise hardcoded
    if site_tab_mapping:
        filters = _build_filters_from_csv(site_tab_mapping, pn, ol, df_all.index)
    else:
        filters = {
            "Charlotte": pn.str.startswith("Charlotte") & ~pn.str.contains("Care Center"),
            "Charlotte Shelter": pn.str.startswith("Charlotte") & pn.str.contains("Care Center"),
            "Citrus": ol.str.contains("Citrus", case=False, na=False),
            "FOX": pn.str.startswith("All-County-VA-Suicide") | pn.str.startswith("Bob Woodruff"),
            "GPD": pn.str.startswith("Pre-Housing") | pn.str.startswith("Retention"),
            "Lake": ol.str.contains("Lake", case=False, na=False),
            "Orlando": pn.str.startswith("Orlando"),
            "Pasco": pn.str.startswith("Pasco"),
            "Pinellas": pn.str.startswith("Pinellas"),
            "Polk": pn.str.startswith("Polk"),
            "PSH": pn.str.contains("PSH", case=False, na=False),
            "San Juan": pn.str.startswith("San Juan"),
            "Sarasota": pn.str.startswith("Sarasota"),
            "Sebring": pn.str.startswith("Sebring"),
            "SouthWest": pn.str.startswith("SouthWest"),
            "Tampa": pn.str.startswith("Tampa"),
        }

    filters["All"] = pd.Series(True, index=df_all.index)

    tabs = OrderedDict()

    for tab_name in SITE_TAB_ORDER:
        mask = filters[tab_name]
        tab_df = df_all[mask].copy()
        tab_df.sort_values(by="Program Name", inplace=True)
        tab_df.reset_index(drop=True, inplace=True)

        if tab_name == "FOX":
            # Reduced 12-column layout — drop 7 columns
            drop_cols = [c for c in FOX_DROP_COLUMNS if c in tab_df.columns]
            tab_df = tab_df.drop(columns=drop_cols)

        elif tab_name in LOCATION_TABS:
            # Add Location column at position 9 (column 10 in 1-indexed)
            loc_values = office_location.reindex(df_all[mask].index).reset_index(drop=True)
            tab_df.insert(9, "Location", loc_values)

        tabs[tab_name] = tab_df
        logger.info("  %-20s %d rows, %d columns", tab_name, len(tab_df), len(tab_df.columns))

    # Warn about programs that don't match any site filter (only on All tab)
    matched = pd.Series(False, index=df_all.index)
    for name, mask in filters.items():
        if name != "All":
            matched = matched | mask
    unmatched = df_all[~matched]
    if len(unmatched) > 0:
        unmatched_programs = unmatched["Program Name"].unique()
        logger.warning(
            "%d program(s) appear only on the All tab (no site match): %s",
            len(unmatched_programs),
            list(unmatched_programs),
        )

    return tabs


def apply_formatting(wb):
    """Apply visual formatting to the output workbook.

    Args:
        wb: openpyxl Workbook object (already written with data)
    """
    for ws in wb.worksheets:
        # Format header row
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-filter on header row
        if ws.max_column and ws.max_row > 1:
            last_col_letter = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        # Build a header-name-to-column-index map for targeted formatting
        header_map = {}
        for col_idx, cell in enumerate(ws[1], 1):
            header_map[str(cell.value or "")] = col_idx

        # Auto-fit column widths and format date/center/red columns
        for col_idx, col_cells in enumerate(ws.columns, 1):
            col_cells = list(col_cells)
            header_text = str(col_cells[0].value or "")

            # Calculate width
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            max_len = max(max_len, 8)  # minimum 8 characters
            if header_text == "Program Name":
                max_len = max(max_len, 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

            # Format date columns
            if header_text in DATE_COLUMNS:
                for cell in col_cells[1:]:
                    if cell.value is not None:
                        cell.number_format = DATE_FORMAT

            # Center-align specific columns
            if header_text in CENTER_COLUMNS:
                for cell in col_cells[1:]:
                    cell.alignment = Alignment(horizontal="center")

            # Red fill for Days With no Service/Contact > 30
            if header_text in RED_THRESHOLD_COLUMNS:
                for cell in col_cells[1:]:
                    try:
                        if cell.value is not None and float(cell.value) > RED_THRESHOLD_DAYS:
                            cell.fill = RED_FILL
                            cell.font = RED_FONT
                    except (ValueError, TypeError):
                        pass

        # Alternating row shading (skip cells already marked red)
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    if cell.fill != RED_FILL:
                        cell.fill = ALT_ROW_FILL


# ---------------------------------------------------------------------------
# MAIN ENTRY POINT
# ---------------------------------------------------------------------------


def main(data_report_card=None, legal_referral=None, output_path=None):
    """Run the full Caseload Report pipeline.

    Args:
        data_report_card: Path to Data Report Card .xlsx (auto-detected if None)
        legal_referral: Path to Legal Services Referral .xlsx (auto-detected if None)
        output_path: Output file path (default: Current_Caseload_{date}.xlsx)
    """
    # Resolve output folder relative to script directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, OUTPUT_DIR)
    os.makedirs(output_dir, exist_ok=True)

    if output_path is None:
        output_path = os.path.join(output_dir, f"Current_Caseload_{date.today().isoformat()}.xlsx")
    elif not os.path.isabs(output_path):
        output_path = os.path.join(output_dir, output_path)

    # Auto-detect input files from folders if not specified
    if data_report_card is None:
        data_report_card = find_xlsx_in_folder(INPUT_DIR_DATA_REPORT_CARD)
    if legal_referral is None:
        legal_referral = find_xlsx_in_folder(INPUT_DIR_LEGAL_REFERRAL)

    # Load input files
    drc, office_location = load_data_report_card(data_report_card)
    legal = load_legal_referral(legal_referral)

    # Load config files
    staff_roster = load_staff_roster()
    site_tab_mapping = load_site_tab_mapping()
    program_validation = load_program_validation()

    # Process the main sheet
    logger.info("Processing main sheet...")
    df_all, office_loc_aligned = process_main_sheet(
        drc, office_location, legal, staff_roster, program_validation
    )

    # Create site tabs
    logger.info("Creating site tabs...")
    tabs = create_site_tabs(df_all, office_loc_aligned, site_tab_mapping)

    # Write to Excel
    logger.info("Writing output to %s", output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for tab_name, tab_df in tabs.items():
            tab_df.to_excel(writer, sheet_name=tab_name, index=False)

    # Reopen for formatting
    logger.info("Applying formatting...")
    wb = load_workbook(output_path)
    apply_formatting(wb)
    wb.save(output_path)

    logger.info("Done. Output: %s", output_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate SVdP CARES Caseload Report from CaseWorthy exports.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Input folders (auto-detected):\n"
            "  input/data_report_card/    — Place Data Report Card .xlsx here\n"
            "  input/legal_referral/      — Place Legal Services Referral .xlsx here\n"
            "\n"
            "Examples:\n"
            "  python caseload_report.py\n"
            "  python caseload_report.py -o My_Report.xlsx\n"
            "  python caseload_report.py --drc report.xlsx --legal referral.xlsx\n"
        ),
    )
    parser.add_argument(
        "--drc",
        default=None,
        help="Path to Data Report Card .xlsx (default: auto-detect from input/data_report_card/)",
    )
    parser.add_argument(
        "--legal",
        default=None,
        help="Path to Legal Services Referral .xlsx (default: auto-detect from input/legal_referral/)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Output file path (default: Current_Caseload_{YYYY-MM-DD}.xlsx)",
    )
    args = parser.parse_args()

    main(
        data_report_card=args.drc,
        legal_referral=args.legal,
        output_path=args.output,
    )
